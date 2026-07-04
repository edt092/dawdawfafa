"""Generación de reportes Excel/PDF de entidad y contratista (plan Pro).

Excel primero por ser más simple (openpyxl, sin dependencias de renderizado
de página); PDF con reportlab en un formato simple — texto y tablas, sin
gráficas embebidas, para no complicar el MVP (ver scalability.md).

Ambos formatos comparten la misma estructura de datos (`ReportData`): KPIs +
una tabla secundaria (top contratistas/entidades) + una terciaria (evolución
mensual / distribución por estado) + una lista de contratos recientes. Así
entity_report_data/contractor_report_data alimentan un solo par de
funciones de render en vez de duplicar la construcción del documento.
"""

import re
from datetime import date
from io import BytesIO
from typing import TypedDict

from fastapi import HTTPException
from openpyxl import Workbook
from openpyxl.styles import Font
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy import extract, func, select
from sqlalchemy.orm import Session

from src.load.models import Contract, Entity, Supplier

TOP_N = 10
RECENT_N = 20


class ReportData(TypedDict):
    titulo: str
    kpis: list[tuple[str, str]]
    secondary_title: str
    secondary_rows: list[tuple[str, float]]
    tertiary_title: str
    tertiary_headers: list[str]
    tertiary_rows: list[tuple]
    recent_title: str
    recent_headers: list[str]
    recent_rows: list[tuple]


def safe_filename(nombre: str) -> str:
    return re.sub(r"[^A-Za-z0-9_-]+", "_", nombre).strip("_")[:80] or "reporte"


def entity_report_data(db: Session, nombre: str) -> ReportData:
    entity = db.execute(select(Entity).where(Entity.nombre_canonico == nombre)).scalars().first()
    if not entity:
        raise HTTPException(status_code=404, detail="Entidad no encontrada.")

    kpis = db.execute(
        select(
            func.count(Contract.id).label("total"),
            func.coalesce(func.sum(Contract.valor), 0).label("valor_total"),
            func.count(Contract.supplier_id.distinct()).label("contratistas_unicos"),
        ).where(Contract.entity_id == entity.id)
    ).mappings().one()

    top_contratistas = db.execute(
        select(Supplier.nombre.label("nombre"), func.sum(Contract.valor).label("valor_total"))
        .join(Contract, Contract.supplier_id == Supplier.id)
        .where(Contract.entity_id == entity.id)
        .group_by(Supplier.nombre)
        .order_by(func.sum(Contract.valor).desc())
        .limit(TOP_N)
    ).mappings().all()

    evolucion = db.execute(
        select(
            extract("year", Contract.fecha).label("anio"),
            extract("month", Contract.fecha).label("mes"),
            func.count(Contract.id).label("cantidad"),
            func.sum(Contract.valor).label("valor_total"),
        )
        .where(Contract.entity_id == entity.id)
        .group_by("anio", "mes")
        .order_by("anio", "mes")
    ).mappings().all()

    recientes = db.execute(
        select(Supplier.nombre.label("contratista"), Contract.valor, Contract.fecha, Contract.estado)
        .join(Supplier, Contract.supplier_id == Supplier.id)
        .where(Contract.entity_id == entity.id)
        .order_by(Contract.fecha.desc())
        .limit(RECENT_N)
    ).mappings().all()

    return ReportData(
        titulo=f"Reporte de entidad — {entity.nombre_canonico}",
        kpis=[
            ("Total contratos", f"{kpis['total']:,}"),
            ("Valor total", f"${float(kpis['valor_total']):,.0f} COP"),
            ("Contratistas únicos", f"{kpis['contratistas_unicos']:,}"),
        ],
        secondary_title="Principales contratistas",
        secondary_rows=[(r["nombre"], float(r["valor_total"])) for r in top_contratistas],
        tertiary_title="Evolución mensual",
        tertiary_headers=["Periodo", "Cantidad", "Valor total (COP)"],
        tertiary_rows=[
            (f"{int(r['anio'])}-{int(r['mes']):02d}", r["cantidad"], float(r["valor_total"]))
            for r in evolucion
        ],
        recent_title="Últimos contratos",
        recent_headers=["Contratista", "Valor (COP)", "Fecha", "Estado"],
        recent_rows=[
            (r["contratista"], float(r["valor"]), _fmt_date(r["fecha"]), r["estado"] or "")
            for r in recientes
        ],
    )


def contractor_report_data(db: Session, nombre: str) -> ReportData:
    supplier = db.execute(select(Supplier).where(Supplier.nombre == nombre)).scalars().first()
    if not supplier:
        raise HTTPException(status_code=404, detail="Contratista no encontrado.")

    kpis = db.execute(
        select(
            func.count(Contract.id).label("total"),
            func.coalesce(func.sum(Contract.valor), 0).label("valor_total"),
            func.count(Contract.entity_id.distinct()).label("entidades_unicas"),
        ).where(Contract.supplier_id == supplier.id)
    ).mappings().one()

    top_entidades = db.execute(
        select(Entity.nombre_canonico.label("nombre"), func.sum(Contract.valor).label("valor_total"))
        .join(Contract, Contract.entity_id == Entity.id)
        .where(Contract.supplier_id == supplier.id)
        .group_by(Entity.nombre_canonico)
        .order_by(func.sum(Contract.valor).desc())
        .limit(TOP_N)
    ).mappings().all()

    por_estado = db.execute(
        select(Contract.estado, func.count(Contract.id).label("cantidad"))
        .where(Contract.supplier_id == supplier.id, Contract.estado.isnot(None))
        .group_by(Contract.estado)
        .order_by(func.count(Contract.id).desc())
    ).mappings().all()

    recientes = db.execute(
        select(Entity.nombre_canonico.label("entidad"), Contract.valor, Contract.fecha, Contract.estado)
        .join(Entity, Contract.entity_id == Entity.id)
        .where(Contract.supplier_id == supplier.id)
        .order_by(Contract.fecha.desc())
        .limit(RECENT_N)
    ).mappings().all()

    return ReportData(
        titulo=f"Reporte de contratista — {supplier.nombre}",
        kpis=[
            ("Total contratos", f"{kpis['total']:,}"),
            ("Valor total", f"${float(kpis['valor_total']):,.0f} COP"),
            ("Entidades cliente", f"{kpis['entidades_unicas']:,}"),
        ],
        secondary_title="Entidades cliente",
        secondary_rows=[(r["nombre"], float(r["valor_total"])) for r in top_entidades],
        tertiary_title="Distribución por estado",
        tertiary_headers=["Estado", "Cantidad"],
        tertiary_rows=[(r["estado"], r["cantidad"]) for r in por_estado],
        recent_title="Contratos recientes",
        recent_headers=["Entidad", "Valor (COP)", "Fecha", "Estado"],
        recent_rows=[
            (r["entidad"], float(r["valor"]), _fmt_date(r["fecha"]), r["estado"] or "")
            for r in recientes
        ],
    )


def _fmt_date(d: date | None) -> str:
    return d.isoformat() if d else ""


def build_workbook(report: ReportData) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen"
    bold = Font(bold=True)

    ws.append([report["titulo"]])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])

    ws.append(["KPI", "Valor"])
    for cell in ws[ws.max_row]:
        cell.font = bold
    for label, value in report["kpis"]:
        ws.append([label, value])

    ws.append([])
    ws.append([report["secondary_title"]])
    ws.cell(ws.max_row, 1).font = Font(bold=True, size=12)
    ws.append(["Nombre", "Valor total (COP)"])
    for cell in ws[ws.max_row]:
        cell.font = bold
    for nombre, valor in report["secondary_rows"]:
        ws.append([nombre, valor])

    ws.append([])
    ws.append([report["tertiary_title"]])
    ws.cell(ws.max_row, 1).font = Font(bold=True, size=12)
    ws.append(report["tertiary_headers"])
    for cell in ws[ws.max_row]:
        cell.font = bold
    for row in report["tertiary_rows"]:
        ws.append(list(row))

    ws2 = wb.create_sheet(report["recent_title"][:31])
    ws2.append(report["recent_headers"])
    for cell in ws2[1]:
        cell.font = bold
    for row in report["recent_rows"]:
        ws2.append(list(row))

    for sheet in wb.worksheets:
        for col in sheet.columns:
            length = max((len(str(c.value)) for c in col if c.value is not None), default=10)
            sheet.column_dimensions[col[0].column_letter].width = min(max(length + 2, 10), 50)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _styled_table(data: list[list]) -> Table:
    t = Table(data, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1D4ED8")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F1F5F9")]),
    ]))
    return t


def build_pdf(report: ReportData) -> BytesIO:
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter, topMargin=1.5 * cm, bottomMargin=1.5 * cm)
    styles = getSampleStyleSheet()
    story = [Paragraph(report["titulo"], styles["Title"]), Spacer(1, 12)]

    kpi_data = [["KPI", "Valor"]] + [[k, v] for k, v in report["kpis"]]
    story += [_styled_table(kpi_data), Spacer(1, 16)]

    story.append(Paragraph(report["secondary_title"], styles["Heading2"]))
    sec_data = [["Nombre", "Valor total (COP)"]] + [
        [n, f"{v:,.0f}"] for n, v in report["secondary_rows"]
    ]
    story += [_styled_table(sec_data), Spacer(1, 16)]

    story.append(Paragraph(report["tertiary_title"], styles["Heading2"]))
    ter_data = [report["tertiary_headers"]] + [list(map(str, row)) for row in report["tertiary_rows"]]
    story += [_styled_table(ter_data), Spacer(1, 16)]

    story.append(Paragraph(report["recent_title"], styles["Heading2"]))
    rec_data = [report["recent_headers"]] + [list(map(str, row)) for row in report["recent_rows"]]
    story.append(_styled_table(rec_data))

    doc.build(story)
    buf.seek(0)
    return buf
